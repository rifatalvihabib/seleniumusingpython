import openpyxl
path="D:\Seleniumbasics\dataread.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active
#sheet=workbook.get_sheet_by_name("sheet1") #multiple sheet
rows = sheet.max_row
columns = sheet.max_column
print(rows)
print(columns)
for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(row=r, column=c).value,end="    ")


    print()